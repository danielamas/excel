from openpyxl import Workbook
from openpyxl import load_workbook
import os

def write_rows(sheet, data_row):
    if sheet and len(data_row) > 0:
        for row in data_row:
            sheet.append(row)

def create_sheet(sheet_title, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = ("T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10")
    write_header(ws, headers)

    data_row = []
    for row in range(1,1000001):
        data_column = []
        for value in range(1, 11):
            data_column.append("Valor{}".format(value))
        data_row.append(data_column)
    write_rows(ws, data_row)

    print(wb.sheetnames)
    wb.save(file_name)

def check_first_row_is_empty(ws, headers, start_row, start_column):
    if ws and len(headers) > 0:
        insert_row = False
        for col in range(0, len(headers)):
            if ws.cell(row=start_row, column=start_column).value != None or ws.cell(row=start_row, column=start_column).value != "":
                insert_row = True
                break
            start_column += 1

        if insert_row:
            ws.insert_rows(0)

# row must be a list or tuple
def write_row(ws, row):
    if ws and len(row) > 0:
        for row in row:
            ws.append(row)

def create_big_sheet(file_path):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    col = ['%d' % i for i in range(6)]
    for irow in range(2500000):
        ws.append(col)

    wb.save(file_path)

# Always first line
def write_header(ws, headers):
    start_row = 1
    start_column = 1
    check_first_row_is_empty(ws, headers, start_row, start_column)
    if ws and len(headers) > 0:
        for h in headers:
            ws.cell(row=start_row, column=start_column, value=h)
            start_column += 1


def open_sheet(path):
    wb = load_workbook(path)
    return (wb.active, wb)

def main():
    # sheet_title = "consolidacao de alarmes"
    # file_name = "teste.xlsx"
    # create_sheet(sheet_title, file_name)

    path = "C:{}tmp{}big_file.xlsx".format(os.sep, os.sep)
    create_big_sheet(path)

    if os.path.isfile(path):
        ws, wb = open_sheet(path)
        headers = ['T%d' % i for i in range(1,7)]
        write_header(ws, headers)
        wb.save(path)
    else:
        print("File {} not exist".format(path))

if (__name__ == "__main__"):
    main()

